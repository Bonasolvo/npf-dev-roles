import re
import os
import json
import shutil
import urllib
import zipfile
import tempfile
from io import BytesIO
from zipfile import ZipFile

from django.template import Context, Template as DjangoTemplate
from docx import Document
from django.http import HttpResponse
from django.core import serializers
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from npf.core.printform.models import Template
from npf.core.xmin.settings import XMIN_ADMINFORM_SERIALIZER, XMIN_INLINE_ADMINFORMS_SERIALIZER


def get_obj_data(context, request):
    """
    Сериализация и добавление формы в JSON ответ
    """
    adminform_serializer = serializers.get_serializer(XMIN_ADMINFORM_SERIALIZER)()
    inline_adminforms_serializer = serializers.get_serializer(XMIN_INLINE_ADMINFORMS_SERIALIZER)()
    response = HttpResponse(content_type='application/json')
    response.write('{')
    response.write('"success": true, ')
    response.write('"perms": {')
    response.write('}, ')
    response.write('"adminform":')
    adminform_serializer.serialize(context['adminform'], ensure_ascii=False, stream=response)
    response.write(', ')
    response.write('"adminform_inline":')
    inline_adminforms_serializer.serialize(context['inline_admin_formsets'],
                                           request=request, ensure_ascii=False, stream=response)
    response.write('}')
    return response


def cleanhtml(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext


class CreateDoc():
    def __init__(self, request, model, model_admin):
        self.request = request
        self.model = model
        self.model_admin = model_admin

    def _create_fields_data(self, id):
        result = self.model_admin.change_view(self.request, str(id))
        result = get_obj_data(result.context_data, self.request)
        result = result.serialize()
        result = result.decode("utf-8")
        result = result.replace('Content-Type: application/json\r\n\r\n', '')
        result = self._find_values('data', result)
        result = self._merge_values(result)
        return result

    def _merge_values(self, values):
        result = {}
        for i in values:
            if isinstance(i, dict):
                result.update(i)
            if isinstance(i, list):
                for sub in i:
                    result.update(sub)
        return result

    def _find_values(self, key, json_repr):

        results = []

        def _decode_dict(a_dict):
            try:
                results.append(a_dict[key])
            except KeyError:
                pass
            return a_dict

        json.loads(json_repr, object_hook=_decode_dict)
        return results

    def _replace_tags(self, folder_path, internal_path_txt, fields_data):
        """
        Замена тегов [[...]] в файле. Подстановка данных.
        """
        with open(folder_path + internal_path_txt, encoding='utf-8') as f:
            text = f.read()

            new_text = text

            pattern = re.compile("\[\[(.*?)\]\]")
            tags = pattern.findall(text)

            # Замена простых тегов
            for tag in tags:
                if cleanhtml(tag) in fields_data:
                    new_text = new_text.replace('[[' + tag + ']]', str(fields_data[cleanhtml(tag)]))

            # Обработка тегов циклов [foreach|endforeach]
            forpattern = re.compile("\[\[.*?foreach(.*?)\]\](.*?)\[\[.*?endforeach.*?\]\]",
                                    re.MULTILINE)
            forvpattern = re.compile("\(\((.*?)\)\)", re.MULTILINE)
            fors = forpattern.findall(new_text)
            for vr, cn in fors:
                if cleanhtml(vr) in fields_data:
                    out = ''
                    for tt in fields_data[cleanhtml(vr)]:
                        tmpout = cn
                        vars = forvpattern.findall(cn)
                        for v in vars:
                            tmpout += tmpout.replace('((' + v + '))', str(tt[cleanhtml(v)]))
                        out += tmpout
                if out != '':
                    new_text = re.sub(
                        "\[\[[^\[]*?foreach" + vr + "\]\]" + cn + "\[\[.*?endif.*?\]\]",
                        out, new_text
                    )

            # Обработка тегов условий [if|else|endif]
            pattern = re.compile("\[\[.*?if(.*?)\]\](.*?)(\[\[.*?endif.*?\]\])", re.MULTILINE)
            epattern = re.compile("\[\[.*?else.*?\]\](.*?)\[", re.MULTILINE)
            ifs = pattern.findall(new_text)
            for i, fi, el in ifs:
                cond1 = cleanhtml(i).strip()
                cond1 = cond1.split('|')
                state2 = False
                for cond2 in cond1:
                    cond2 = cond2.split('&amp;')
                    state1 = True
                    for cond in cond2:
                        state = False
                        cond = cond.strip()
                        # a = b
                        if cond.find('=') != -1:
                            part_cont = cond.split('=')
                            if part_cont[0].find('`') != -1:
                                if part_cont[0].replace('`', '') == fields_data[part_cont[1]]:
                                    state = True
                            elif part_cont[1].find('`') != -1:
                                if part_cont[1].replace('`', '') == fields_data[part_cont[0]]:
                                    state = True
                            else:
                                if fields_data[part_cont[0]] != fields_data[part_cont[1]]:
                                    state = True
                        # a > b
                        elif cond.find('>') != -1:
                            part_cont = cond.split('>')
                            if part_cont[0].find('`') != -1:
                                if part_cont[0].replace('`', '') > fields_data[part_cont[1]]:
                                    state = True
                            elif part_cont[1].find('`') != -1:
                                if part_cont[1].replace('`', '') > fields_data[part_cont[0]]:
                                    state = True
                            else:
                                if fields_data[part_cont[0]] > fields_data[part_cont[1]]:
                                    state = True
                        # a < b
                        elif cond.find('<') != -1:
                            part_cont = cond.split('<')
                            if part_cont[0].find('`') != -1:
                                if part_cont[0].replace('`', '') < fields_data[part_cont[1]]:
                                    state = True
                            elif part_cont[1].find('`') != -1:
                                if part_cont[1].replace('`', '') < fields_data[part_cont[0]]:
                                    state = True
                            else:
                                if fields_data[part_cont[0]] < fields_data[part_cont[1]]:
                                    state = True
                        if not state:
                            state1 = False
                    if state1:
                        state2 = True

                ift = fi
                elses = epattern.findall(el)

                if elses:
                    elset = elses[0]
                else:
                    elset = ''

                if not state2:
                    new_text = re.sub("\[\[[^\[]*?if(.*?)\]\]" + fi + "\[\[.*?endif.*?\]\]",
                                      elset, new_text)
                else:
                    new_text = re.sub("\[\[[^\[]*?if(.*?)\]\]" + fi + "\[\[.*?endif.*?\]\]",
                                      ift, new_text)

        os.remove(folder_path + internal_path_txt)
        with open(folder_path + internal_path_txt, 'w', encoding='utf-8') as f:
            f.write(new_text)

    def _replace_tags_name(self, name, fields_data):
        """
            Замена тегов [[...]] в имени файла
        """
        pattern = re.compile("\[\[(.*?)\]\]")
        tags = pattern.findall(name)
        new_text = name
        for tag in tags:
            if cleanhtml(tag) in fields_data:
                new_text = new_text.replace('[[' + tag + ']]', str(fields_data[cleanhtml(tag)]))

        return new_text

    def _compile_doc(self, template_path, fields_data, doc_name):
        """
            Разархивация docx файла, преобразование и обратная архивация.
        """
        doc = Document(template_path)
        doc.save(doc_name)

        docx_path = doc_name
        zip_path = docx_path.replace('docx', 'zip')
        folder_path = docx_path.replace('.docx', '')
        internal_path_xml = '/word/document.xml'
        internal_path_txt = '/word/document.txt'

        # docx > zip
        os.rename(docx_path, zip_path)
        # unzip
        zipfile.ZipFile(zip_path).extractall(path=folder_path)
        # remove initial zip
        os.remove(zip_path)
        # xml > txt
        os.rename(folder_path + internal_path_xml, folder_path + internal_path_txt)
        # replace_tags
        self._replace_tags(folder_path, internal_path_txt, fields_data)
        # txt > xml
        os.rename(folder_path + internal_path_txt, folder_path + internal_path_xml)
        # zip
        shutil.make_archive(folder_path, 'zip', folder_path)
        # remove folder
        shutil.rmtree(folder_path)
        # zip > docx
        os.rename(zip_path, docx_path)

    def run(self):
        files = {}
        ids = json.loads(self.request.GET.get('ids'))
        self._create_fields_data(ids[0])
        template_name = self.request.GET.get('doc_name')
        template_path = Template.objects.get(name=template_name).file
        file_name = Template.objects.get(name=template_name).file_name
        for id in ids:
            fields_data = self._create_fields_data(id)
            file_path = tempfile.mktemp(suffix='.docx')
            self._compile_doc(template_path, fields_data, file_path)
            files[self._replace_tags_name(str(id) + '_' + file_name, fields_data)] = file_path

        return files


def create_excel(data, module_name, columns, excel_heading):
    def _create_heading_cells(column_manager, columns_to_delete):
        class ProcessNode():
            """
            Класс для того, чтобы на основе входной древовидной структуры шапки таблицы
            сделать матрицу (m*n) ячеек для записи их в файл
            """
            def __init__(self, node):
                self.node = node
                self.result = []
                self.parents_width = {}
                self.max_depth = 0
                self.node_layout = []

            def _walk_node(self, node, depth, parents, first=False):
                if depth > self.max_depth:
                    self.max_depth = depth
                if 'columns' in node:
                    first = True
                    for item in node['columns']:
                        self._walk_node(item, depth + 1, parents[:] + [node['text']], first=first)
                        first = False
                self.result.append({
                    'text': node.get('text'),
                    'depth': depth,
                    'parent': parents[-1] if parents else None
                })
                if first:
                    _parents = [parents[-1]]
                else:
                    _parents = parents[:]
                for parent in _parents:
                    if parent not in self.parents_width:
                        self.parents_width[parent] = 1
                    else:
                        self.parents_width[parent] += 1

            def _extend_upper(self):
                for item in self.node_layout:
                    item.append(item[-1])

            def _prepare_node_layout(self):
                for level in range(self.max_depth + 1):
                    level_layout = []
                    first = True
                    for item in self.result:
                        if item['depth'] == level:
                            level_layout.append(item['text'])
                            if not first and item['parent']:
                                self._extend_upper()
                            first = False
                    try:
                        if len(level_layout) < len(self.node_layout[-1]):
                            index = len(self.node_layout[-1]) - len(level_layout)
                            level_layout = self.node_layout[-1][:index] + level_layout
                    except IndexError:
                        pass
                    self.node_layout.append(level_layout)

            def run(self):
                self._walk_node(self.node, 0, [])
                self._prepare_node_layout()
                return self.node_layout

        def _equalize_nodes(node_layouts, max_depth):
            for node in node_layouts:
                shortage = max_depth - len(node)
                if shortage:
                    for i in range(shortage):
                        node.append(node[-1])
            return node_layouts

        def _make_result(node_layouts, max_depth):
            result = []
            for i in range(max_depth):
                row_result = []
                for item in node_layouts:
                    try:
                        row_result += item[i]
                    except IndexError:
                        row_result += ['fake']
                result.append(row_result)
            return result

        def _delete_nodes(node_layouts, columns_to_delete):
            if columns_to_delete:
                # mark nodes to delete
                counter = -1
                for item in node_layouts:
                    for i in range(len(item[0])):
                        counter += 1
                        if counter in columns_to_delete:
                            for sub in item:
                                sub[i] += '_delete'
                # delete nodes
                for item in node_layouts:
                    for sub in item:
                        for _sub in sub[:]:
                            if '_delete' in _sub:
                                sub.remove(_sub)

            return node_layouts

        def _get_max_depth(node_layouts):
            max_depth = 0
            for item in node_layouts:
                current_depth = 0
                for sub in item:
                    if sub:
                        current_depth += 1
                max_depth = max(max_depth, current_depth)

            return max_depth

        node_layouts = []
        for node in column_manager:
            node_layout = ProcessNode(node).run()
            node_layouts.append(node_layout)

        node_layouts = _delete_nodes(node_layouts, columns_to_delete)
        max_depth = _get_max_depth(node_layouts)
        node_layouts = _equalize_nodes(node_layouts, max_depth)
        result = _make_result(node_layouts, max_depth)
        return result

    def _make_raw_heading(cells):
        for row in cells:
            ws.append(row)
        for row in ws.rows:
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=True,
                    indent=0
                )

    def _make_merge_columns(cells):
        merge_columns = []
        for row in cells:
            merge_cell_row = []
            start, end = 0, 0
            while True:
                while end <= len(row) - 1 and row[start] == row[end]:
                    end += 1
                merge_cell_row.append((start, end - 1))
                if end == len(row):
                    merge_columns.append(merge_cell_row)
                    break
                start = end
        return merge_columns

    def _make_set_of_merge_columns(merge_columns):
        _set = []
        for row in merge_columns:
            _set += row
        _set = list(set(_set))
        return _set

    def _make_cells_to_merge(_set, merge_columns):
        _cells_to_merge = []
        for cell in _set:
            positions = []
            for row in merge_columns:
                if cell in row:
                    positions.append(merge_columns.index(row))
            _cells_to_merge.append({
                'columns': cell,
                'rows': (min(positions), max(positions))
            })
        return _cells_to_merge

    def _merge_cells(cells_to_merge):
        for item in cells_to_merge:
            ws.merge_cells(
                start_row=item['rows'][0] + 1,
                start_column=item['columns'][0] + 1,
                end_row=item['rows'][1] + 1,
                end_column=item['columns'][1] + 1
            )

    def _create_heading(columns_to_delete):
        cells = _create_heading_cells(excel_heading, columns_to_delete)
        _make_raw_heading(cells)
        merge_columns = _make_merge_columns(cells)
        _set = _make_set_of_merge_columns(merge_columns)
        cells_to_merge = _make_cells_to_merge(_set, merge_columns)
        _merge_cells(cells_to_merge)

    def _write_data():
        for item in data['data']:
            row = []
            for i in columns:
                if i['visible']:
                    try:
                        row.append(item[i['name']])
                    except KeyError:
                        row.append('---')
                        pass
            ws.append(row)

    def _render_to_xlsx(xlsx, context):
        tmp = BytesIO()
        with ZipFile(tmp, 'w') as document_zip, ZipFile(xlsx) as template_zip:
            template_archive = {name: template_zip.read(name) for name in template_zip.namelist()}
            template_xml = template_archive.pop('xl/workbook.xml')
            for n, f in template_archive.items():
                document_zip.writestr(n, f)
            t = DjangoTemplate(template_xml)
            document_zip.writestr('xl/workbook.xml', t.render(Context(context)))
        return tmp

    def _columns_to_delete():
        columns_to_delete = []
        counter = -1
        for item in columns:
            counter += 1
            if not item['visible']:
                columns_to_delete.append(counter)
        return columns_to_delete

    # создаём книгу
    wb = Workbook()
    ws = wb.active

    # определяем неотображаемые колонки
    columns_to_delete = _columns_to_delete()
    # создаём и записываем (в файл) шапку
    _create_heading(columns_to_delete)
    # записываем (в файл) данные
    _write_data()

    # сохраняем файл и готовим его к отправке
    file_name = module_name + '.xlsx'
    file_path = tempfile.mktemp(suffix='.xlsx')

    wb.save(file_path)
    zp = _render_to_xlsx(file_path, {})

    # формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(
        urllib.parse.quote_plus(file_name).replace('+', ' '))
    response.write(zp.getvalue())

    os.remove(file_path)

    return response
